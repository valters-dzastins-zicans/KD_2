import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver import ActionChains
from datetime import date
from datetime import datetime

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)


workbook = load_workbook('AtrastāsMašīnas.xlsx') 
sheet = workbook.active
current_row = sheet.max_row + 1

today = date.today()
sodiena = today.strftime("%m/%Y")
print("Today's date:", sodiena)


        
url = "https://www.ss.com/lv/transport/cars/search/"
driver.get(url)

time.sleep(2)
cenaNo = driver.find_element(By.NAME, value="topt[8][min]")
cenaLidz = driver.find_element(By.NAME, value="topt[8][max]")

TehniskaOpt = driver.find_element(By.NAME, value="opt[223][]")
TehniskaOpt.click()

print("Ievadiet mašīnas cenu no")
cn = input()
print("Ievadiet mašīnas cenu līdz")
cl = input()
cenaNo.send_keys(cn)
cenaLidz.send_keys(cl)
print("Ievadiet vēlamo tehniskās apskates ilgumu (mēnešos)")
men = input()
menTeh = int(men)

option = driver.find_element(By.XPATH, "//option[@value='Yes']")
option.click()
option.click()

meklet = driver.find_element(By.XPATH, "//input[@value='Meklēt']")
time.sleep(1)
meklet.click()
time.sleep(2)

scroll_origin = ScrollOrigin.from_viewport(10, 10)
ActionChains(driver)\
    .scroll_from_origin(scroll_origin, 0, 99999)\
    .perform()
time.sleep(2)




masinas = driver.find_elements(By.XPATH, "//tr[starts-with(@id, 'tr_')]")
masina_count = len(masinas)
time.sleep(1)

page_number = 2

while True:
    try:
        
        button = driver.find_element(By.XPATH, "//a[contains(@href, 'page" + str(page_number) + ".html')]")
        

       
        for i in range(masina_count-1):
            try:
                time.sleep(1)
                current_masinas = driver.find_elements(By.XPATH, "//tr[starts-with(@id, 'tr_')]")
                time.sleep(1)
                current_masinas[i].click()
                time.sleep(1)

                tehniska = driver.find_element(By.ID, "tdo_223")
                print(tehniska.text)
                datums = datetime.strptime(tehniska.text, "%m.%Y").date()
                if datums.year == 2023:
                    difference = 0
                elif datums.year > 2024:
                    difference = 999
                    
                else:
                    difference = datums.month - today.month

                
                if difference >= menTeh: 
                    print("OK")
                    marka = driver.find_element(By.ID, "tdo_31")
                    markaText = marka.text
                    gads = driver.find_element(By.ID, "tdo_18")
                    gadsText = gads.text
                    motors = driver.find_element(By.ID, "tdo_15")
                    motorsText = motors.text
                    nobraukums = driver.find_element(By.ID, "tdo_16")
                    nobraukumsText = nobraukums.text
                    teha = driver.find_element(By.ID, "tdo_223")
                    tehaText = teha.text 
                    
                    
                    links = driver.current_url

                    sheet.cell(row=current_row, column=1, value=markaText)
                    sheet.cell(row=current_row, column=2, value=gadsText)
                    sheet.cell(row=current_row, column=3, value=motorsText)
                    sheet.cell(row=current_row, column=4, value=nobraukumsText)
                    sheet.cell(row=current_row, column=5, value=tehaText)
                    sheet.cell(row=current_row, column=6, value=links)
                    

                    
                    current_row += 1
    
                driver.back()

                scroll_origin = ScrollOrigin.from_viewport(10, 10)
                ActionChains(driver)\
                .scroll_from_origin(scroll_origin, 0, 200)\
                .perform()


            except NoSuchElementException:
                print("Tehniskā nav ierakstīta")
                driver.back()
                continue

        
        button = driver.find_element(By.XPATH, "//a[contains(@href, 'page" + str(page_number) + ".html')]")
        button.click()
        page_number += 1

    except NoSuchElementException:
        print("Nav papildus lapu")
        for i in range(masina_count-1):
            try:
                time.sleep(1)
                current_masinas = driver.find_elements(By.XPATH, "//tr[starts-with(@id, 'tr_')]")
                current_masinas[i].click()
                time.sleep(1)

                tehniska = driver.find_element(By.ID, "tdo_223")
                print(tehniska.text)
                datums = datetime.strptime(tehniska.text, "%m.%Y").date()
                if datums.year == 2023:
                    difference = 0
                elif datums.year > 2024:
                    
                    difference = 999
                else:
                    difference = datums.month - today.month

                
                if difference >= menTeh: 
                    print("OK")          
                    marka = driver.find_element(By.ID, "tdo_31")
                    markaText = marka.text
                    gads = driver.find_element(By.ID, "tdo_18")
                    gadsText = gads.text
                    motors = driver.find_element(By.ID, "tdo_15")
                    motorsText = motors.text
                    nobraukums = driver.find_element(By.ID, "tdo_16")
                    nobraukumsText = nobraukums.text
                    teha = driver.find_element(By.ID, "tdo_223")
                    tehaText = teha.text 
                    
                    
                    links = driver.current_url

                    sheet.cell(row=current_row, column=1, value=markaText)
                    sheet.cell(row=current_row, column=2, value=gadsText)
                    sheet.cell(row=current_row, column=3, value=motorsText)
                    sheet.cell(row=current_row, column=4, value=nobraukumsText)
                    sheet.cell(row=current_row, column=5, value=tehaText)
                    sheet.cell(row=current_row, column=6, value=links)
                    

                    
                    current_row += 1
                driver.back()

                scroll_origin = ScrollOrigin.from_viewport(10, 10)
                ActionChains(driver)\
                .scroll_from_origin(scroll_origin, 0, 200)\
                .perform()


            except NoSuchElementException:
                print("Tehniskā nav ierakstīta")
                driver.back()
                continue
        break        
        
saglabat = input("Vai vēlaties saglabāt excel file? (Y/N) ") 
if saglabat == "Y":
    workbook.save('AtrastāsMašīnas.xlsx')
    driver.quit()
else:
    driver.quit()



